import re, sys, os, pytz, yaml, markdown, asyncio, inspect
import pandas as pd
import xml.etree.ElementTree as ET

from typing import List, Dict, Tuple, Union, Literal
from pathlib import Path
from pydantic import BaseModel, Field
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from docx.shared import Pt, Mm
from docx.enum.section import WD_ORIENT
from htmldocx import HtmlToDocx
from bs4 import BeautifulSoup
from docx import Document

from core.config import settings

# ----------------------------------------------------------------------------------------------------
# 0. 경로 설정
# ----------------------------------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parents[4]

# 입력 파일 경로
REPORT_ROOT = BASE_DIR / "agent" / "report_generation" / "reports"
YAML_PATH = REPORT_ROOT / "config" / "report.yaml"
RESOURCE_PATH = REPORT_ROOT / "resources"
PROMPT_PATH = RESOURCE_PATH / "prompt"
SQL_PATH = RESOURCE_PATH / "sql"
REPORT_TEMPLATE_PATH = RESOURCE_PATH / "report_template"

# 결과 파일 경로
REPORT_OUTPUT_PATH = settings.REPORT_OUTPUT_PATH
OUTPUT_DIR = BASE_DIR / REPORT_OUTPUT_PATH
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

with open(YAML_PATH, encoding="utf-8") as f:
    CFG = yaml.safe_load(f)
REPORT_NAME = CFG["briefing"]["reports"]["report_name"]
CATEGORIES = CFG["briefing"]["reports"]["categories"]

AZURE_OPENAI_API_KEY = settings.AZURE_OPENAI_API_KEY

# ----------------------------------------------------------------------------------------------------
# 1. SQL 실행 관련 함수
# ----------------------------------------------------------------------------------------------------
async def _get_executor() -> "MCPToolExecutor":
    """
    전역 _executor 가 None 이면 get_mcp_executor()로 초기화
    이미 초기화돼 있으면 그대로 반환

    Returns:
        MCPToolExecutor
    """
    global _executor
    if _executor is None:
        async with _executor_lock:         # 동시에 두 개가 초기화되는 것을 방지
            if _executor is None:          # 두 번째 체크 (double‑checked locking)
                _executor = await get_mcp_executor()
    return _executor

async def SQL_executor(
    mcp_executor,
    SQL: str,
    db_name: str
) -> pd.DataFrame:
    """
    MCP executor로 SQL 실행 -> DataFrame 반환

    Args:
        mcp_executor  : MCPToolExecutor
        SQL     (str) : 실행할 SQL
        db_name (str) : SQL을 실행할 대상 DataBase 식별자

    Returns:
        pd.DataFrame : SQL 실행 결과
    """
    res = await mcp_executor.execute_tool(db_name, {'query': SQL})
    extracted = [row for row in res]
    df = pd.DataFrame(extracted)
    return df

async def SQL2df(
    mcp_executor,
    sql_path: str,
    sql_file: str,
    df_name: str,
    db_name: str,
    params: dict
) -> Tuple[str, pd.DataFrame]:
    """
    SQL 파일경로, DB, f-string 변수 입력하여 DataFrame으로 변환하는 함수

    Args:
        mcp_executor    : MCPToolExecutor
        sql_path (str)  : SQL 파일이 위치한 디렉터리 경로
        sql_file (str)  : 실행할 SQL 파일명
        df_name  (str)  : 반환 시 함께 제공될 데이터프레임 라벨
        db_name  (str)  : SQL을 실행할 대상 DataBase 식별자
        params   (dict) : SQL 텍스트 내 {key} 형태의 플레이스홀더에 치환할 값 딕셔너리

    Returns:
        Tuple[str, pd.DataFrame]:
            df_name : 반환 시 함께 제공될 데이터프레임 라벨
            df      : SQL 실행 결과
    """
    with open(f'{sql_path}/{sql_file}', 'r', encoding='utf-8') as f:
        SQL = f.read().format(**params)
    df = await SQL_executor(mcp_executor, SQL, db_name)
    return df_name, df

# ----------------------------------------------------------------------------------------------------
# 2. 보고서 생성 관련 함수
# ----------------------------------------------------------------------------------------------------
# --------------------------------------------------
# 2-0) 공통
# --------------------------------------------------
def replace_placeholders(
    text: str,
    mapping: dict
) -> str:
    """
    {placeholder} 를 mapping 값으로 교체

    Args:
        text    (str)  : 치환할 대상 문자열
        mapping (dict) : 치환에 사용할 매핑 딕셔너리 (키는 "{...}" 형태의 전체 placeholder)

    Returns:
        str : 매핑에 따라 placeholder가 치환된 문자열
    """
    pattern = re.compile(r'\{([^}]+)\}')
    return pattern.sub(lambda m: mapping.get(m.group(0), m.group(0)), text)

def df_to_xml(
    df: pd.DataFrame,
    root_name: str = 'data',
    row_name: str = 'row'
) -> str:
    """
    pd.DataFrame를 XML 문자열로 변환

    Args:
    df        (pd.DataFrame) : 변환할 데이터프레임
    root_name (str)          : 최상위 루트 요소 이름
    row_name  (str)          : 각 행을 감싸는 요소 이름

    Returns:
        str : utf-8 인코딩된 XML 문자열
    """
    root = ET.Element(root_name)
    for _, r in df.iterrows():
        row_el = ET.SubElement(root, row_name)
        for col, val in r.items():
            cell = ET.SubElement(row_el, str(col))
            cell.text = '' if pd.isna(val) else str(round(float(val), 1)) if is_number(val) else str(val)
    return ET.tostring(root, encoding='utf-8', xml_declaration=True).decode('utf-8')

def is_number(v: str) -> bool:
    """
    입력값이 소수형인지 판별

    Args:
        v (str) : 문자열(숫자 변환이 가능할 경우에 대한 처리)

    Returns:
        bool : 변환 가능하면 True, ValueError가 발생하면 False
    """
    try:
        float(v)
        return True
    except ValueError:
        return False

# --------------------------------------------------
# 2-1) Word 형식 보고서 관련 함수
# --------------------------------------------------
def replace_channel_column(cat: str, mapping: Dict[str, str]) -> Dict[str, str]:
    """
    주어진 카테고리(cat)에 따라 mapping 딕셔너리의
    {spec_col} 과 {spec_name} 값을 설정한다.
    """
    if cat == "브랜드":
        mapping["{spec_col}"] = "이용거래기간내용"
        mapping["{spec_name}"] = "거래기간"
    elif cat == "플랫폼":
        mapping["{spec_col}"] = "플랫폼이용빈도내용"
        mapping["{spec_name}"] = "최근 1개월내 이용빈도별"
    elif cat == "대면채널":
        mapping["{spec_col}"] = "영업점이용빈도내용"
        mapping["{spec_name}"] = "최근 3개월 내 방문횟수별"
    elif cat == "고객센터":
        mapping["{spec_col}"] = "고객센터이용빈도내용"
        mapping["{spec_name}"] = "최근 6개월 내 이용빈도별"
    elif cat == "상품":
        mapping["{spec_col}"] = "고객경험단계명"
        mapping["{spec_name}"] = "상품군별"
    return mapping

# 2-1-1) SQL 준비 (템플릿 + placeholder)
def load_sql_template(SQL_PATH, query_id: str) -> str:
    sql_path = SQL_PATH / f"{query_id}.sql"
    if not sql_path.is_file():
        return None
    try:
        return sql_path.read_text(encoding="utf-8")
    except Exception as e:
        return None

def prepare_sql(SQL_PATH, BASE_PLACEHOLDER_MAP, query_id: str, cat: str, extra_map: Dict[str, str] | None = None) -> str:
    """query_id → 최종 실행 SQL 문자열"""
    tpl = load_sql_template(SQL_PATH, query_id)
    if tpl is None:                     # 파일이 없을 경우
        return None

    # placeholder 매핑 구성
    mapping: Dict[str, str] = BASE_PLACEHOLDER_MAP.copy()
    mapping["{channel}"] = cat                      # 채널(카테고리) 정보
    mapping = replace_channel_column(cat, mapping)  # 채널별 spec_col, spec_name 컬럼

    if extra_map:
        mapping.update(extra_map)

    return replace_placeholders(tpl, mapping)

# 2-1-2️) 쿼리 실행 (단일·다중 모두 지원)
async def run_query(
    SQL_PATH: str,
    BASE_PLACEHOLDER_MAP: dict,
    query_id: str | List[str],
    cat: str,
    mcp_executor,
    extra_map: Dict[str, str] | None = None,
) -> pd.DataFrame:
    """
    query_id 가 리스트이면 순차 실행 후 concat. -> 미정
    SQL 파일이 없으면 빈 DataFrame 반환.
    """
    async def _exec_one(qid: str) -> pd.DataFrame:
        try:
            sql = prepare_sql(SQL_PATH, BASE_PLACEHOLDER_MAP, qid, cat, extra_map)
            if sql is None:                     # 파일이 없을 경우
                return pd.DataFrame()           # 빈 DataFrame
            return await SQL_executor(mcp_executor, sql, 'mysql_query')
        except Exception as e:
            # 여기서 잡히는 예외는 파일 입출력 외에 DB 실행 오류 등
            return pd.DataFrame()               # 오류가 나도 빈 DataFrame 반환

    if isinstance(query_id, list):
        dfs = [await _exec_one(q) for q in query_id]
        return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
    else:
        return await _exec_one(query_id)

# 2-1-3) 전체 query_id → (cat, depth, chunk) 매핑 테이블 생성
def build_query_index(CATEGORIES) -> Dict[Tuple[str, str], Tuple[str, str]]:
    """
    CATEGORIES 구조를 순회하면서
    1. query_id 가 리스트인 경우 각 쿼리마다 별도 chunk 를 만들어 CATEGORIES 를 동적으로 확장
    2. 각 쿼리마다 대응되는 prompt_id 를 보존
    3. (cat, query_id) → (depth, chunk) 를 반환
    """
    idx: Dict[Tuple[str, str], Tuple[str, str]] = {}

    for cat, depths in CATEGORIES.items():
        for depth, chunks in depths.items():
            # depth 자체가 query_id 를 포함하는 경우 (예: 은행 → NPS현황)
            if isinstance(chunks, dict) and "query_id" in chunks:
                qids = chunks["query_id"]
                prompt_ids = chunks.get("prompt_id")

                # 단일 쿼리
                if not isinstance(qids, list):
                    idx[(cat, qids)] = (depth, "root")
                    continue

                # 리스트인 경우, 하나의 chunk 안에 여러 query_id 를 보관
                for i, qid in enumerate(qids):
                    idx[(cat, qid)] = (depth, "root")  # chunk 이름은 그대로
                continue

            # 일반 청크 구조 (예: 브랜드 → NPS진단 → 청크1, 청크2 …)
            # 청크 이름을 순회할 때는 딕셔너리 변형을 방지하기 위해 keys() 구분자
            for chunk_name, info in list(chunks.items()):
                qids = info["query_id"]
                prompt_id = info.get("prompt_id")

                # 단일 query_id
                if not isinstance(qids, list):
                    idx[(cat, qids)] = (depth, chunk_name)
                    continue

                # 리스트인 경우, 원본 dict는 그대로 두고 하나의 chunk 안에 여러 query_id 를 보관
                for i, qid in enumerate(qids):
                    idx[(cat, qid)] = (depth, chunk_name)

    return idx

# 2-1-4️) 전체 쿼리 비동기 실행
async def fetch_all_queries(SQL_PATH, BASE_PLACEHOLDER_MAP, QUERY_INDEX, mcp_executor) -> Dict[Tuple[str, str], pd.DataFrame]:
    """
    반환 형태: {(cat, query_id): DataFrame}
    """
    tasks: Dict[Tuple[str, str], asyncio.Task] = {}
    for (cat, qid), (depth, chunk) in QUERY_INDEX.items():
        # extra placeholder 에 현재 카테고리 전달
        tasks[(cat, qid)] = asyncio.create_task(
            run_query(SQL_PATH, BASE_PLACEHOLDER_MAP, qid, cat, mcp_executor, extra_map={"channel": cat})
        )

    results: Dict[Tuple[str, str], pd.DataFrame] = {}
    for key, task in tasks.items():
        results[key] = await task
    return results

# 2-1-5️) 후처리 함수 정의 (쿼리‑별)
def apply_postprocess(
    POSTPROCESS_MAP: dict,
    cat: str,
    qid: str,
    df: Union[pd.DataFrame, List[pd.DataFrame]],
    bm_info: pd.DataFrame | None,
) -> Union[pd.DataFrame, List[pd.DataFrame]]:
    """
    df 가 List[DataFrame] 인 경우 각 요소에 대해 개별 후처리를 수행.
    """
    if isinstance(df, list):
        return [apply_postprocess(cat, qid, d, bm_info) for d in df]

    if qid not in POSTPROCESS_MAP:
        return df

    fn, need_bm, extra = POSTPROCESS_MAP[qid]
    extra = extra.copy()
    if "channel" in extra:
        extra["channel"] = cat

    if need_bm:
        if bm_info is None:
            return df
        try:
            return fn(df, bm_info, **extra)   # type: ignore[arg-type]
        except Exception as e:
            return df
    else:
        try:
            return fn(df)
        except Exception as e:
            return df

# 2-1-5‑1) TD_은행_01  (NPS 현황)
def postprocess_TD_은행_01(df: pd.DataFrame) -> pd.DataFrame:
    """
    - 반환값: 후처리된 DataFrame
    """
    if df.empty:
        return df
    return df

# 2-1-5‑2) TD_은행_02  (NPS 영향요인)
def postprocess_TD_은행_02(df: pd.DataFrame) -> pd.DataFrame:
    """
    - 벤치마크사 컬럼을 동일하게 변환
    """
    if df.empty:
        return df
    return df

# 2-1-5‑3) TD_은행_03  (고객불만지수)
def postprocess_TD_은행_03(df: pd.DataFrame, bm_info: pd.DataFrame) -> pd.DataFrame:
    """
    - 거래은행구분 변환 후 bm_info와 inner join
    """
    if df.empty:
        return df
    # bm_info 에는 이미 채널구분, 거래은행구분, NPS순위 등이 들어있음
    df = pd.merge(bm_info, df, on=['채널구분', '거래은행구분'], how='inner')
    return df

# 2-1-5‑4) 채널(브랜드/플랫폼/대면채널/고객센터/상품) – NPS 진단 청크 1
def postprocess_채널_청크1(df: pd.DataFrame) -> pd.DataFrame:
    """
    """
    if df.empty:
        return df
    return df

# 2-1-5‑5) 채널 청크2 (TD_채널_01_2) – NPS 진단
def postprocess_채널_청크2(df: pd.DataFrame, bm_info: pd.DataFrame, channel: str) -> pd.DataFrame:
    """
    - 거래은행구분 변환 후 bm_info(해당 channel)와 merge
    """
    if df.empty:
        return df
    df = pd.merge(df, bm_info[bm_info['채널구분'] == channel], on=['거래은행구분', '채널구분'])
    return df

# 2-1-5‑6) 채널 청크3 (TD_채널_01_3) – NPS 진단
postprocess_채널_청크3 = postprocess_채널_청크2

# 2-1-5‑7) 채널 청크4 – NPS 진단 (TD_채널_01_4) – 로직은 청크2와 동일
postprocess_채널_청크4 = postprocess_채널_청크2

# 2-1-5‑8) 채널 청크2 – 서비스품질진단 (TD_채널_02_2) – rename + merge
def postprocess_채널_서비스품질_청크2(df: pd.DataFrame, bm_info: pd.DataFrame, channel: str) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.rename(columns={'영향요인구분': '서비스품질요소'})
    df = pd.merge(df, bm_info[bm_info['채널구분'] == channel], on=['거래은행구분', '채널구분'])
    return df

# 2-1-5‑9) 채널 청크1 – 서비스품질진단 (TD_채널_02_1) – 변환만
postprocess_채널_서비스품질_청크1 = postprocess_채널_청크2  # 동일 로직

# 2-1-5‑10) 채널 청크1 – 고객불만지수 (TD_채널_03_1) – 변환 + merge
postprocess_채널_고객불만_청크1 = postprocess_채널_청크2

POSTPROCESS_MAP: dict[str, tuple[callable, bool, dict]] = {
    # ── 은행 ───────────────────────────────────────────────────────
    "TD_은행_01": (postprocess_TD_은행_01, False, {}),          # NPS 현황
    "TD_은행_02": (postprocess_TD_은행_02, False, {}),          # NPS 영향요인
    "TD_은행_03": (postprocess_TD_은행_03, True,  {}),          # 고객불만지수 (bm_info 필요)

    # ── 채널(브랜드/플랫폼/대면채널/고객센터/상품) ───────────────────────
    # 청크1 – NPS 진단
    "TD_채널_01_1": (postprocess_채널_청크1, False, {}),

    # 청크2 – NPS 진단 (bm_info + channel 전달)
    "TD_채널_01_2": (postprocess_채널_청크2, True,
                     {"channel": None}),   # channel 은 실행 시 채워짐

    "TD_채널_01_3": (postprocess_채널_청크3, True,
                     {"channel": None}),

    # 청크4 – NPS 진단 (청크2와 동일)
    "TD_채널_01_4": (postprocess_채널_청크4, True,
                     {"channel": None}),

    # 청크2 – 서비스품질진단
    "TD_채널_02_2": (postprocess_채널_서비스품질_청크2, True,
                     {"channel": None}),

    # 청크1 – 서비스품질진단 (청크2와 동일 로직)
    "TD_채널_02_1": (postprocess_채널_서비스품질_청크1, True,
                     {"channel": None}),

    # 청크1 – 고객불만지수 (청크2와 동일 로직)
    "TD_채널_03_1": (postprocess_채널_고객불만_청크1, True,
                     {"channel": None}),
}

# 2-1-6) 은행 전용 후처리 (bm_info, pstv_fac, weak_fac, weak1_fac)
def postprocess_bank(df_nps: pd.DataFrame, df_factor: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    기존 로직을 유지하면서, 앞서 정의한 `postprocess_TD_은행_01/02` 를 적용한다.
    """
    # ① NPS 현황 후처리
    df_nps = postprocess_TD_은행_01(df_nps)
    # ② factor(영향요인) 후처리
    df_factor = postprocess_TD_은행_02(df_factor)

    # ---- bm_info -------------------------------------------------
    df_nps["BM_KEY"] = df_nps["NPS순위"] - 1
    df_nps = pd.merge(df_nps[['채널구분', '거래은행구분', 'NPS순위']], df_nps[df_nps['거래은행구분'] == 'KB국민은행'][['채널구분', 'BM_KEY']], on='채널구분')
    bm_info = df_nps[(df_nps['NPS순위'] == 1) | (df_nps['거래은행구분'] == 'KB국민은행') | (df_nps['NPS순위'] == df_nps['BM_KEY'])].drop(columns='BM_KEY')

    # ---- factor (pstv / weak / weak1) ---------------------------
    df_noch = df_factor[df_factor['영향요인구분'] != '채널전체'].reset_index(drop=True)
    df_noch['영향도'] = df_noch['영향도'].astype(float)
    df_noch['벤치마크사_영향도'] = df_noch['벤치마크사_영향도'].astype(float)

    # 긍정영향요인 (각 채널당 영향도 1위)
    pos_idx = df_noch.groupby('채널구분')['영향도'].idxmax().values
    pstv_fac = df_noch.loc[pos_idx][['채널구분', '영향요인구분']].rename(columns={'영향요인구분': '긍정영향요인'})

    # 약점영향요인 (GAP < -0.11, 긍정 제외)
    df_noch['영향도GAP'] = df_noch['영향도'] - df_noch['벤치마크사_영향도']
    weak_idx = list(set(df_noch[df_noch['영향도GAP'] < -0.11].index) - set(pos_idx))
    weak_fac = df_noch.loc[weak_idx][['채널구분', '영향요인구분']].rename(columns={'영향요인구분': '약점영향요인'})
    
    # 최약점 (채널당 GAP 최솟값)
    weak1_idx = df_noch.groupby('채널구분')['영향도GAP'].idxmin().values
    weak1_fac = df_noch.loc[weak1_idx][['채널구분', '영향요인구분']].rename(columns={'영향요인구분': '약점영향요인'})

    return bm_info, pstv_fac, weak_fac, weak1_fac

# 2-1-7) LLM 호출 – depth/chunk 별 보고서 생성
async def generate_report_for_chunk(
    CATEGORIES,
    BASE_PLACEHOLDER_MAP,
    user_id,
    llm,
    OUTPUT_DIR,
    prompts: dict,
    cat: str,
    depth: str,
    chunk: str,
    df: Union[pd.DataFrame, List[pd.DataFrame]],
    placeholders: Dict[str, str],
) -> Tuple[str, str]:
    """
    * `placeholders` 에는 {bm_info}, {pstv_fac} 등 필요 시 삽입할 markdown 문자열이 들어 있다.
    * prompt 파일은 yaml 에 정의된 `prompt_id` 로부터 읽는다.
    * -------------------------------------------------------------------------------------
    * `df` 가 단일 DataFrame 이거나, query_id 가 리스트일 때 여러 DataFrame 으로 전달될 수 있다.
    * 리스트인 경우 내부적으로 concat 으로 하나의 DataFrame 으로 결합한다.
    """
    # root 청크인지 확인하고 prompt_id 추출
    if chunk == "root":
        prompt_id = CATEGORIES[cat][depth]["prompt_id"]
    else:
        prompt_id = CATEGORIES[cat][depth][chunk]["prompt_id"]

    prompt = prompts.get(prompt_id + ".txt", "")

    # placeholder 삽입
    for key, md in placeholders.items():
        prompt = prompt.replace(key, md)

    # 데이터 삽입
    if isinstance(df, list):
        # 리스트 -> {data1}, {data2}
        for i, d in enumerate(df, start=1):
            placeholder = f"{{data{i}}}"
            prompt = prompt.replace(placeholder, df_to_xml(d))
        if "{data}" in prompt:
            all_md = "\n".join(df_to_xml(d) for d in df)
            prompt = prompt.replace("{data}", all_md)
    
    else:
        prompt = prompt.replace("{data}", df_to_xml(df))

    # 채널별 spec_col, spec_name 삽입
    mapping: Dict[str, str] = BASE_PLACEHOLDER_MAP.copy()
    mapping["{channel}"] = cat
    mapping = replace_channel_column(cat, mapping)
    prompt = replace_placeholders(prompt, mapping)

    # LLM 호출
    result_md = await llm.ainvoke(prompt)

    # 파일 저장 (markdown)
    out_md_path = OUTPUT_DIR / REPORT_NAME / f"{cat}_{depth}_{chunk}.md"
    out_md_path.parent.mkdir(parents=True, exist_ok=True)
    out_md_path.write_text(result_md.content, encoding="utf-8")

    return cat, result_md.content

# 2-1-8) 카테고리별 파일 저장 함수
async def save_category_reports(BASE_PLACEHOLDER_MAP, REPORT_NAME, CATEGORY_MAP, OUTPUT_DIR, cat_md_map: Dict[str, List[str]], combine_with_bank: bool) -> None:
    """
    각 카테고리별 markdown을 하나로 합치고, 필요하면 docx 로 변환한다.
    - 저장 위치:  output/<REPORT_NAME>/category/<cat>_report.md
    - docx:   output/<REPORT_NAME>/category/<cat>_report.docx
    """
    yyyy = BASE_PLACEHOLDER_MAP["{yyyy}"]
    yyyyhf = BASE_PLACEHOLDER_MAP["{yyyyhf}"]
    cat_dir = OUTPUT_DIR
    cat_dir.mkdir(parents=True, exist_ok=True)

    for cat, parts in cat_md_map.items():
        category = CATEGORY_MAP.get(cat, "")
        bank_md_path = cat_dir / f"{REPORT_NAME}_{yyyy}_{yyyyhf}_bank.md"
        if (combine_with_bank and
            cat != '은행' and
            bank_md_path.exists()):
            bank_md = bank_md_path.read_text(encoding="utf-8")

            cat_md_path = cat_dir / f"{REPORT_NAME}_{yyyy}_{yyyyhf}_{category}.md"
            cat_docx_path = cat_dir / f"{REPORT_NAME}_{yyyy}_{yyyyhf}_{category}.docx"
            cat_md = cat_md_path.read_text(encoding="utf-8")
            com_cat_md = f"{bank_md}\n\n{cat_md}"
            cat_md_path.write_text(com_cat_md, encoding="utf-8")
            md_to_docx(cat_md_path, cat_docx_path)

        else:
            # 순서는 `QUERY_INDEX`대로 
            cat_md = "\n".join(parts).strip()

            # 파일명 만들기 → td_nps_report_2025_하반기_은행.md
            file_name = f"{REPORT_NAME}_{yyyy}_{yyyyhf}_{category}.md"

            md_path = cat_dir / file_name
            md_path.write_text(cat_md, encoding="utf-8")

            file_name_docx = f"{REPORT_NAME}_{yyyy}_{yyyyhf}_{category}.docx"
            docx_path = cat_dir / file_name_docx
            md_to_docx(OUTPUT_DIR / file_name, OUTPUT_DIR / file_name_docx)

def set_heading_styles(doc):
    """
    python-docx Document 객체의 기본 텍스트 및 Heading 1~3 스타일 일괄 설정

    Args:
        doc (docx.document.Document): 스타일을 적용할 대상 문서 객체

    Effects:
        - Normal: 폰트 Arial, 크기 10pt, 굵게 해제
        - Heading 1: 폰트 Arial, 크기 20pt, 굵게
        - Heading 2: 폰트 Arial, 크기 16pt, 굵게 해제
        - Heading 3: 폰트 Arial, 크기 12pt, 굵게 해제

    Returns:
        None (전달된 문서 객체의 스타일을 수정)

    Raises:
        KeyError: 문서에 해당 스타일('Normal', 'Heading 1', 'Heading 2', 'Heading 3')이 없을 경우
    """
    normal = doc.styles['Normal']
    normal.font.size = Pt(10)
    normal.font.bold = False
    normal.font.name = 'Arial'

    h1 = doc.styles['Heading 1']
    h1.font.size = Pt(20)
    h1.font.bold = True
    h1.font.name = 'Arial'

    h2 = doc.styles['Heading 2']
    h2.font.size = Pt(16)
    h2.font.bold = False
    h2.font.name = 'Arial'

    h3 = doc.styles['Heading 3']
    h3.font.size = Pt(12)
    h3.font.bold = False
    h3.font.name = 'Arial'

def md_to_docx(
    md_path: Path,
    docx_path: Path
) -> None:
    """
    markdown -> docx (이미지 경로 보정 포함)

    Args:
        md_path   (pathlib.Path) : 변환할 Markdown 파일 경로
        docx_path (pathlib.Path) : 생성될 docx 파일 경로

    Returns:
        None (파일 생성)
    """
    if not md_path.is_file():
        raise FileNotFoundError(md_path)

    base_dir = md_path.parent
    md_text = md_path.read_text(encoding="utf-8")
    html = markdown.markdown(md_text, extensions=["tables"])

    soup = BeautifulSoup(html, "html.parser")
    for img in soup.find_all("img"):
        src = img.get("src")
        if src and not src.startswith(("http://", "https://")):
            if not os.path.isabs(src):
                abs_path = os.path.normpath(os.path.join(base_dir, src))
                if os.path.exists(abs_path):
                    img["src"] = abs_path

    doc = Document()
    section = doc.sections[0]
    section.orientation = WD_ORIENT.LANDSCAPE
    section.page_width, section.page_height = section.page_height, section.page_width
    section.top_margin = Mm(15)
    section.bottom_margin = Mm(15)
    section.left_margin = Mm(15)
    section.right_margin = Mm(15)

    parser = HtmlToDocx()
    parser.add_html_to_document(str(soup), doc)

    set_heading_styles(doc)

    for table in doc.tables:
        table.style = "Table Grid"

    doc.save(docx_path)

# --------------------------------------------------
# 2-2) Excel 형식 보고서 관련 함수
# --------------------------------------------------
def replace_match(
    v: str,
    val_dict: dict
) -> str:
    """
    문자열 내의 템플릿 플레이스홀더를 딕셔너리 값으로 치환
    플레이스홀더 형식은 {{key}}
    딕셔너리에 해당 key가 없으면 원래 플레이스홀더를 그대로 유지
    치환되는 값은 str()로 문자열 변환되어 삽입

    Args:
        v        (str)  : 치환 대상이 되는 원본 문자열
        val_dict (dict) : key-value 매핑을 가진 딕셔너리

    Returns:
        str : 플레이스홀더가 가능한 범위에서 치환된 문자열
    """
    pattern = re.compile(r'\{\{(\w+)\}\}')
    def replacer(match):
        key = match.group(1)
        return str(val_dict.get(key, match.group(0)))
    return pattern.sub(replacer, v)

def append_val_to_dict(
    df: pd.DataFrame,
    val_dict: dict
) -> dict:
    """
    DataFrame의 각 요소를 Dictionary에 incremental하게 추가하는 함수
    Args:
        df       (pd.DataFrame) : 키를 생성할 기준이 되는 pd.DataFrame
        val_dict (dict)         : 결과를 누적할 대상 dictionary

    Returns:
        dict : 입력된 val_dict에 새로운 키-값 쌍이 추가된 동일 딕셔너리 객체
    """
    for idx, row in df.iterrows():
        for col in df.columns:
            val = str(row[col])
            val_dict[f'{col}_{idx+1}'] = val
    return val_dict

class BU_Complaint(BaseModel):
    complaint_summary: str = ''
    voc_example1: str = ''
    voc_example2: str = ''
    voc_example3: str = ''

class RG_Complaint(BaseModel):
    CX: str = ''
    complaint_summary: str = ''
    voc_example1: str = ''
    voc_example2: str = ''
    voc_example3: str = ''

class BU_Complaints(BaseModel):
    complaint1: BU_Complaint
    complaint2: BU_Complaint
    complaint3: BU_Complaint

class RG_Complaints(BaseModel):
    complaint1: RG_Complaint
    complaint2: RG_Complaint
    complaint3: RG_Complaint

async def run_llm(
    prompt_path: str,
    prompt_file: str,
    user_id,
    llm,
    params: dict,
    gen_type: Literal['default', 'BU_VOC', 'RG_VOC'] = 'default'
):
    """
    프롬프트를 load한 후, parameter를 포맷한 뒤 비동기 LLM 호출을 수행
    JSON 형식의 출력을 기대할 경우 Pydantic 모델로 검증한 후 pd.DataFrame으로 반환

    Args:
        prompt_path (str) : 프롬프트 템플릿(.txt) 파일이 위치한 디렉터리 경로
        prompt_file (str) : 사용할 프롬프트 템플릿 파일명
        user_id           :
        llm               : 
        params     (dict) : 프롬프트 템플릿 내부의 플레이스홀더를 대체할 dictionary
        gen_type    (str) : 생성 유형 지정 (default: 자연어 생성, BU_VOC: BU VOC 요약, RG_VOC: 지역영업그룹 VOC 요약)

    Returns:
        gen_type:
            default        -> LLM 응답 TEXT로 반환
            BU_VOC, RG_VOC -> BaseModel 형태에 맞게 JSON으로 반환
    """

    gen_type_dict = {
        'BU_VOC': BU_Complaints,
        'RG_VOC': RG_Complaints,
    }

    prompt = Path(f'{prompt_path}/{prompt_file}.txt').read_text(encoding='utf-8').format(**params)
    response = await llm.ainvoke(prompt)
    if gen_type != 'default':
        Complaint = gen_type_dict.get(gen_type)
        try:
            output_valid = Complaint.model_validate_json(response.content)
            return pd.DataFrame(output_valid.model_dump())
        except Exception as e:
            return None
    return response.content

def replace_placeholders_in_workbook(
    input_path: str,
    output_path: str,
    val_dict: dict[str, str]
) -> int:
    """
    Excel 파일의 모든 시트의 변수를 Dictionary의 Value 값으로 변환하고 저장

    Args:
        input_path  (str) : 보고서 템플릿 엑셀 파일 경로
        output_path (str) : 템플릿 기반 생성할 엑셀 파일 경로
        val_dict   (dict) : 플레이스홀더 치환에 사용할 딕셔너리

    Returns:
        int : 치환되어 값이 변경된 셀의 총 개수
    """
    wb = load_workbook(input_path)
    changed = 0
    for ws in wb.worksheets: 
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue

                if '{{' not in v or '}}' not in v:
                    continue

                new_v = replace_match(v, val_dict)
                if new_v == v:
                    cell.value = None
                elif new_v.isdigit() and '.' not in new_v:
                    cell.value = int(new_v)
                    cell.number_format = '#,##0'
                elif is_number(new_v):
                    cell.value = float(new_v)
                    cell.number_format = '#,##0.0;[Red]-#,##0.0'
                elif new_v == 'None':
                    cell.value = ''
                else:
                    cell.value = new_v
                changed += 1

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return changed

async def get_list_region_group(mcp_executor) -> list:
    """
    가장 최신 날짜 기준 지역그룹관리구분 리스트를 생성
    """
    SQL = f"""
    SELECT DISTINCT 지역그룹관리구분명
    FROM INST1.TSCCVCI08
    WHERE 기준년월일 = (
        SELECT MAX(기준년월일)
        FROM INST1.TSCCVCI08
    )
    """
    df = await SQL_executor(mcp_executor, SQL, 'mysql_query')
    list_region_group = df['지역그룹관리구분명'].tolist()
    return list_region_group

# ----------------------------------------------------------------------------------------------------
# 3. job 비동기 처리 함수 -> retrying 방지를 위한 동기 처리로 변경 (2026.01.06)
# ----------------------------------------------------------------------------------------------------
async def run_sql_jobs(
    jobs_info: list,
    mcp_executor,
    SQL_PATH: str
) -> dict[str, pd.DataFrame]:
    """
    보고서 생성을 위한 데이터 조회 및 저장

    Args:
        jobs_info (list) : 실행 job 리스트
            df_name  : df_dict에서 사용할 데이터프레임 라벨
            sql_file : 조회할 SQL 파일명
            db_name  : SQL을 실행할 대상 DataBase 식별자
            params   : SQL 텍스트 내 {key} 형태의 플레이스홀더에 치환할 값 딕셔너리
        mcp_executor     :
        SQL_PATH   (str) : SQL 파일이 위치한 디렉터리 경로

    Returns:
        dict[str, str] : job별 생성 결과를 포함한 dict
    """
    result = [] 
    for job in jobs_info:
        df_name, sql_file, db_name, params = job
        res = await SQL2df(
            mcp_executor,
            SQL_PATH,
            sql_file,
            df_name,
            db_name,
            params
        )
        result.append(res)
    return result

async def make_script_content(
    jobs_info: list,
    df_dict: dict,
    user_id,
    llm,
    PROMPT_PATH: str
):
    """
    보고서에서 데이터 기반의 인사이트 문구를 생성

    Args:
        jobs_info  (list) : 실행 job 리스트
            df          : 인사이트 문구를 생성하기 위해 사용할 데이터
            script_id   : 보고서 템플릿 내 script 변수명
            prompt_file : 사용 프롬프트 파일명
        df_dict    (dict) : job을 수행하기 위해 필요한 데이터가 존재하는 DataFrame 모음 사전
        user_id           : 
        llm               : 
        PROMPT_PATH (str) : 프롬프트 경로

    Returns:
        dict[str, str] : job별 생성 결과를 포함한 dict
    """
    result = {}
    for job in jobs_info:
        df, script_id, prompt_file = job
        res = await run_llm(
            PROMPT_PATH,
            prompt_file,
            user_id,
            llm,
            {'data': df_to_xml(df_dict[df])},
        )
        result[script_id] = res
    return result

async def make_voc_summary(
    jobs_info: list,
    df_dict: dict,
    user_id,
    llm,
    PROMPT_PATH: str,
    gen_type: Literal['BU_VOC', 'RG_VOC']
):
    """
    기준 기간에 수집된 불만 VOC에 대한 요약 및 예시 생성

    Args:
        jobs_info  (list) : 실행 job 리스트
            df                : 현재 기간 불만 VOC 추출을 위해 df_dict에서 사용할 df_name을 의미
            prompt_file       : 사용 프롬프트 파일명
            channel           : 대상 채널
            region_group_name : 대상 지역영업그룹
        df_dict    (dict) : job을 수행하기 위해 필요한 데이터가 존재하는 DataFrame 모음 사전
        user_id           : 
        llm               : 
        PROMPT_PATH (str) : 프롬프트 경로
        gen_type    (str) : 생성 유형 지정

    Returns:
        list[pd.DataFrame] : job별 생성 결과를 포함한 list
    """
    result = []
    if gen_type == 'BU_VOC':
        for job in jobs_info:
            df, prompt_file, channel = job
            res = await run_llm(
                PROMPT_PATH,
                prompt_file,
                user_id,
                llm,
                {
                    'channel': channel,
                    'voc': '|'.join(df_dict[df]['VOC']),
                },
                gen_type
            )
            result.append(res)
    elif gen_type == 'RG_VOC':
        for job in jobs_info:
            df, prompt_file, region_group_name = job
            res = await run_llm(
                PROMPT_PATH,
                prompt_file,
                user_id,
                llm,
                {
                    'region_group_name': region_group_name,
                    'voc': df_to_xml(df_dict[df]),
                },
                gen_type
            )
            result.append(res)
    return result

async def make_voc_compare(
    jobs_info: list,
    voc_summary: list,
    df_dict: dict,
    user_id,
    llm,
    PROMPT_PATH: str
):
    """
    지난 기간 불만 VOC와 현재 기간 불만 VOC의 내용을 비교분석하여 인사이트 문구 생성

    Args:
        jobs_info   (list) : 실행 job 리스트
            df_last     : 저번 기간 불만 VOC 추출을 위해 df_dict에서 사용할 df_name을 의미
            ch_idx      : voc_summary 리스트에서 사용할 값 idx
            prompt_file : 사용 프롬프트 파일명
            channel     : 대상 채널
        voc_summary (list) : 이번 기간 불만 VOC 요약 내용
        df_dict     (dict) : job을 수행하기 위해 필요한 데이터가 존재하는 DataFrame 모음 사전
        user_id            : 
        llm                : 
        PROMPT_PATH  (str) : 프롬프트 경로

    Returns:
        list[str] : job별 생성 결과를 포함한 list
    """
    result = []
    for job in jobs_info:
        df_last, ch_idx, prompt_file, channel = job
        res = await run_llm(
            PROMPT_PATH,
            prompt_file,
            user_id,
            llm,
            {
                'channel': channel,
                'voc_last': '|'.join(df_dict[df_last]['VOC']),
                'voc_summary': df_to_xml(voc_summary[ch_idx].loc[['complaint_summary']]),
            },
        )
        result.append(res)
    return result

# ----------------------------------------------------------------------------------------------------
# 4. 날짜 처리 함수
# ----------------------------------------------------------------------------------------------------
async def make_td_nps_report_date_params(mcp_executor) -> dict:
    """
    Top-down 보고서에 필요한 날짜를 계산하여 반환
    테이블에 적재된 데이터 기준 가장 최신 조사년도와 반기구분 값 활용

    Args:
        mcp_executor: MCPToolExecutor

    Returns:
        dict:
            yyyy        (str) : 최근 TD 조사년도
            yyyyhf      (str) : 최근 TD 조사반기
            yyyy_b1hf   (str) : 전반기 TD 조사년도
            yyyyhf_b1hf (str) : 전반기 TD 조사반기
    """

    SQL = """
    SELECT 조사년도, 반기구분명
    FROM inst1.TSCCVMGC1
    GROUP BY 1, 2
    ORDER BY 1 DESC, 2 DESC
    LIMIT 1
    """
    last_date   = await SQL_executor(mcp_executor, SQL, 'mysql_query')
    yyyy        = str(last_date['조사년도'].values[0])
    yyyyhf      = last_date['반기구분명'].values[0]
    yyyy_b1hf   = str(int(yyyy)-1) if yyyyhf == '상반기' else yyyy
    yyyyhf_b1hf = '하반기' if yyyyhf == '상반기' else '상반기'

    date_params = {
        'yyyy'       : yyyy,
        'yyyyhf'     : yyyyhf,
        'yyyy_b1hf'  : yyyy_b1hf,
        'yyyyhf_b1hf': yyyyhf_b1hf,
    }

    return date_params

async def make_bu_nps_weekly_report_date_params(mcp_executor) -> dict:
    """
    Bottom-up 주간 보고서에 필요한 날짜를 계산하여 반환

    Args:
        mcp_executor: MCPToolExecutor

    Returns:
        dict:
            monday_b01w        (str) : 1주일 전 월요일에 대한 날짜 (yyyymmdd 형태)
            biz_endday_b01w    (str) : 1주일 전 마지막 영업일에 대한 날짜 (yyyymmdd 형태)
            monday_b02w        (str) : 2주일 전 월요일에 대한 날짜 (yyyymmdd 형태)
            biz_endday_b02w    (str) : 2주일 전 마지막 영업일에 대한 날짜 (yyyymmdd 형태)
            biz_endday_b01m    (str) : 1달 전 마지막 영업일에 대한 날짜 (yyyymmdd 형태)
            yyyymmdd_cx_manage (str) : 고객경험 관리 시작일자 (yyyymmdd 형태) / 2025년의 경우 3월부터로 지정되어 있어서 추가한 변수로 추후에는 불필요
    """
    SQL = f"""
    SELECT
          max(기준년월일) as 기준년월일
    FROM INST1.TSCCVMGF4
    """
    near_date = await mcp_executor.execute_tool('mysql_query', {"query": SQL})
    if not isinstance(near_date, list) or len(near_date) == 0:
        raise Exception(f"TSCCVMGF4의 가장 최근 일자 조회 시 오류가 발생했습니다. {near_date}")

    baseymd = datetime.strptime(near_date[0]["기준년월일"], "%Y%m%d")
    yyyy = baseymd.strftime('%Y')

    monday_b01w = (baseymd - timedelta(days=baseymd.weekday() +  7)).strftime('%Y%m%d')    # 1주 전 월요일
    sunday_b01w = (baseymd - timedelta(days=baseymd.weekday() +  1)).strftime('%Y%m%d')    # 1주 전 일요일
    monday_b02w = (baseymd - timedelta(days=baseymd.weekday() + 14)).strftime('%Y%m%d')    # 2주 전 월요일
    sunday_b02w = (baseymd - timedelta(days=baseymd.weekday() +  8)).strftime('%Y%m%d')    # 2주 전 일요일
    endday_b01m = (baseymd.today().replace(day=1) - timedelta(days=1)).strftime('%Y%m%d')  # 1달 전 마지막일

    SQL = f"""
    SELECT
          기준년월일
        , 최근영업년월일
    FROM INST1.TSCCVCI12 /* 경영정보일자기본 */
    WHERE 기준년월일 IN ('{sunday_b01w}', '{sunday_b02w}', '{endday_b01m}')
    """
    biz_df = await SQL_executor(mcp_executor, SQL, 'mysql_query')
    biz_endday_b01w = biz_df[biz_df['기준년월일'] == sunday_b01w]['최근영업년월일'].values[0]
    biz_endday_b02w = biz_df[biz_df['기준년월일'] == sunday_b02w]['최근영업년월일'].values[0]
    biz_endday_b01m = biz_df[biz_df['기준년월일'] == endday_b01m]['최근영업년월일'].values[0]
    yyyymmdd_cx_manage = '20250301' if yyyy == '2025' else f'{yyyy}0101'

    date_params = {
        'monday_b01w': monday_b01w,
        'biz_endday_b01w': biz_endday_b01w,
        'monday_b02w': monday_b02w,
        'biz_endday_b02w': biz_endday_b02w,
        'biz_endday_b01m': biz_endday_b01m,
        'yyyymmdd_cx_manage': yyyymmdd_cx_manage,
    }

    return date_params

def make_bu_nps_monthly_report_date_params() -> dict:
    """
    Bottom-up 월간 보고서에 필요한 날짜를 계산하여 반환

    Returns:
        dict:
            yyyy             (str) : 현재 연도 (yyyy 형태)
            yyyy01           (str) : 현재 연도의 1월 (yyyymm 형태)
            yyyy0101         (str) : 현재 연도의 1월 1일 (yyyymmdd 형태)
            yyyymm_b01m      (str) : 1달 전 날짜 (yyyymm 형태)
            yyyymm_b02m      (str) : 2달 전 날짜 (yyyymm 형태)
            yyyymm_cx_manage (str) : 고객경험 관리 시작월 (yyyymm 형태) / 2025년의 경우 3월부터로 지정되어 있어서 추가한 변수로 추후에는 불필요
    """
    kst = pytz.timezone('Asia/Seoul')
    baseymd = datetime.now(kst)
    yyyy = baseymd.strftime('%Y')
    yyyy01 = baseymd.strftime('%Y') + '01'
    yyyy0101 = baseymd.strftime('%Y') + '0101'
    yyyymm_b01m = (baseymd - relativedelta(months=1)).strftime('%Y%m')
    yyyymm_b02m = (baseymd - relativedelta(months=2)).strftime('%Y%m')
    yyyymmdd_b01m = (baseymd.today().replace(day=1) - timedelta(days=1)).strftime('%Y%m%d')
    yyyymm_cx_manage = '202503' if yyyy == '2025' else yyyy01

    date_params = {
        'yyyy': yyyy,
        'yyyy01': yyyy01,
        'yyyy0101': yyyy0101,
        'yyyymm_b01m': yyyymm_b01m,
        'yyyymm_b02m': yyyymm_b02m,
        'yyyymmdd_b01m': yyyymmdd_b01m,
        'yyyymm_cx_manage': yyyymm_cx_manage,
    }

    return date_params

def make_region_group_nps_biweekly_report_date_params() -> dict:
    """
    지역영업그룹 격주 보고서에 필요한 날짜를 계산하여 반환

    Returns:
        dict:
            yyyy        (str) : 현재 연도 (yyyy 형태)
            yyyy0101    (str) : 현재 연도의 1월 1일 (yyyymmdd 형태)
            monday_b01w (str) : 1주일 전 월요일에 대한 날짜 (yyyymmdd 형태)
            friday_b01w (str) : 1주일 전 금요일에 대한 날짜 (yyyymmdd 형태)
            monday_b02w (str) : 2주일 전 월요일에 대한 날짜 (yyyymmdd 형태)
            friday_b03w (str) : 3주일 전 금요일에 대한 날짜 (yyyymmdd 형태)
            yyyymm_b01m (str) : 지난 달 마지막 일 (yyyymmdd 형태)
    """
    kst = pytz.timezone('Asia/Seoul')
    baseymd = datetime.now(kst)
    yyyy = baseymd.strftime('%Y')
    yyyy0101 = baseymd.strftime('%Y') + '0101'

    monday_b01w = (baseymd - timedelta(days=baseymd.weekday() +  7)).strftime('%Y%m%d')
    friday_b01w = (baseymd - timedelta(days=baseymd.weekday() +  3)).strftime('%Y%m%d')
    monday_b02w = (baseymd - timedelta(days=baseymd.weekday() + 14)).strftime('%Y%m%d')
    friday_b03w = (baseymd - timedelta(days=baseymd.weekday() + 17)).strftime('%Y%m%d')
    yyyymm_b01m = (baseymd.today().replace(day=1) - timedelta(days=1)).strftime('%Y%m%d')

    date_params = {
        'yyyy'       : yyyy,
        'yyyy0101'   : yyyy0101,
        'monday_b01w': monday_b01w,
        'friday_b01w': friday_b01w,
        'monday_b02w': monday_b02w,
        'friday_b03w': friday_b03w,
        'yyyymm_b01m': yyyymm_b01m,
    }

    return date_params